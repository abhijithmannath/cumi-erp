import openpyxl
import re
import datetime
import tkMessageBox
import utils


def read_orders_from_sheet(filename):
	wb = openpyxl.load_workbook(filename)
	sh = wb.active
	orders = []
	for x in range(2, sh.max_row+1):
		order = []
		for y in range(1,6):
			val = sh.cell(row=x,column=y).value
			if y == 2:
				val = datetime.datetime.strptime(val,'%d/%m/%y').strftime('%Y-%m-%d')
			elif y == 4:
				match = re.search(r'(\d+(?:\.\d+)?)\sX\s(\d+(?:\.\d+)?)\sX\s(\d+(?:\.\d+)?)',val)
				val = float(match.group(2))
				order.append(float(match.group(1)))
			order.append(val)
		orders.append(order)
	return orders

def get_bond_system(bond):
	wb = openpyxl.load_workbook('settings/bond_database.xlsx')
	sh = wb.active
	bs= None
	for x in range(2, sh.max_row+1):
		if sh.cell(row=x, column=2).value == bond:
			bs = sh.cell(row=x, column=1).value
			break
	if not bs:
		tkMessageBox.showerror('Bond Not Found','%s was not found in bond_database.xlsx'%bond )
	return bs

def get_bond_params(bond, dia):
	wb = openpyxl.load_workbook('settings/bond_cycle.xlsx')
	if re.match('^VB.+', bond):
		bond = 'VB1'
	bond_system = get_bond_system(bond)
	if not bond_system:
		return
	sheet = re.search(r'^BS([0-9]+)',bond_system).group(1)
	sheet = int(sheet) - 1
	sh = wb.worksheets[sheet]
	firing_cycle = None

	for x in range(2, sh.max_row+1):
		val = sh.cell(row=x, column=1).value

		if utils.check_val_in_range(val, utils.convert_mm_inch(dia)):
			return {'FIRING_CYCLE': str(sh.cell(row=x, column=3).value), 'COOLING_CYCLE':str(sh.cell(row=x, column=4).value), \
			'GREATER_THAN_SHELF': int(sh.cell(row=x, column=2).value), 'SHELF_EXCLUSIONS': str(sh.cell(row=x, column=5).value)}


def get_wheels_per_base(dia, base_size, shelf_no, bond): #base_size small => 0, Big => 1
	if re.match('^VB.+', bond):
		if base_size == 0:
			return 4
		elif base_size == 1:
			return 8

	col = base_size+2
	wb = openpyxl.load_workbook('settings/wheel_vs_shelf.xlsx')
	main_sheet = wb.worksheets[0]
	sh = None
	for x in range(2,main_sheet.max_row+1):
		if main_sheet.cell(row=x, column=1).value == int(shelf_no):
			sh = main_sheet.cell(row=x, column=2).value - 1
			sh = wb.worksheets[sh]
	if sh is None:
		tkMessageBox.showerror('Shelf %s not found'%shelf_no,'Check wheel_vs_shelf.xlsx!')
		return None
	else:
		for x in range(2, sh.max_row+1):
			val = sh.cell(row=x, column=1).value
			if utils.check_val_in_range(val, dia):
				return int(sh.cell(row=x, column=col).value)
	return None
	tkMessageBox.showerror('Error','Range not found please check settings sheet')



