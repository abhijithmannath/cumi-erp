#TODO justify 50 days kiln update


import sqlite3

conn = sqlite3.connect('cumi.db')
print "Opened database successfully";

def add_setting(key, value):
	conn.execute('''INSERT INTO SETTINGS (KEY,VALUE) VALUES('%s','%s')'''%(key,value))

def create_order_table():
	conn.execute('''CREATE TABLE ORDERS
       (ORDER_ID 	  TEXT 	 PRIMARY KEY   NOT NULL,
       O_DIA 	  	  REAL 	  NOT NULL,
       THICKNESS 	  REAL 	  NOT NULL,
       BS             TEXT    NOT NULL,
       SHIP_DATE      TEXT     NOT NULL,
       ENTRY_DATE     TEXT     NOT NULL,
       QTY	 	  	  INT 	  NOT NULL,
       FNSO 		  INT 	  NOT NULL,
       MOULDED 	  	  INT 	  NOT NULL,
       FIRED 		  INT 	  NOT NULL,
       LEFT_TO_SIMULATE	INT	NOT NULL,
       FIRING_CYCLE		TEXT	,
       COOLING_CYCLE	TEXT ,
       GREATER_THAN_SHELF	INT		NOT NULL, 
       SHELF_EXCLUSIONS		TEXT );''')
	conn.execute('CREATE INDEX order_index ON ORDERS (ORDER_ID, SHIP_DATE, FIRING_CYCLE, MOULDED, FNSO, FIRED);')
	print "Table created successfully";

def create_mould_table():
	conn.execute('''CREATE TABLE MOULDING
		(order_id	INT ,
		qty 	INT NOT NULL,
		FOREIGN KEY (order_id) REFERENCES ORDERS (ORDER_ID) );''')

def create_settings_table():
	import datetime

	conn.execute('''CREATE TABLE SETTINGS
		(KEY	TEXT	NOT NULL,
		VALUE	TEXT	NOT NULL);''')
	add_setting('LAST_REAL_KILN_USED','4')
	add_setting('LAST_SIMULATED_KILN_USED','4')
	add_setting('KILN_UPDATE_ON', (datetime.date.today()+datetime.timedelta(days=50)).strftime('%Y-%m-%d') )


def update_kiln_data():
	from sql import SQL
	import openpyxl
	wb = openpyxl.load_workbook('settings/kiln.xlsx')
	sh = wb.active

	with SQL() as sql:
		sql.delete_kiln_data()
		for x in range(2, sh.max_row+1):
			sql.add_kiln_params(sh.cell(row=x,column=1).value, sh.cell(row=x,column=2).value, sh.cell(row=x,column=3).value)


def create_kiln_params_table():
	conn.execute('''CREATE TABLE KILNS
		(KILN		INT	NOT NULL,
		 MAX_RUNS	INT	NOT NULL,
		 SIZE 		INT 	NOT NULL,
		 RUNS_LEFT	INT 	NOT NULL,
		 SIMULATED_RUNS_LEFT INT NOT NULL );''')
	update_kiln_data()
	



if __name__ == '__main__':
	create_order_table()
	create_mould_table()
	create_settings_table()
	create_kiln_params_table()
